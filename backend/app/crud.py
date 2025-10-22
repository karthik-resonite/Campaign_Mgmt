from sqlalchemy.orm import Session
from . import models, schemas
from .utils.auth import hash_password
import csv
from io import StringIO, BytesIO
import openpyxl  # 👈 for xlsx

def create_company(db: Session, company: schemas.CompanyCreate):
    db_company = models.Company(
        name=company.name,
        username=company.username,
        email=company.email,
        phone=company.phone,
        password=hash_password(company.password),  # hashed
        role=company.role
    )
    db.add(db_company)
    db.commit()
    db.refresh(db_company)
    return db_company

def get_company(db: Session, company_id: int):
    return db.query(models.Company).filter(models.Company.id == company_id).first()

def get_all_company(db: Session):
    return db.query(models.Company).all()

def create_agent(db: Session, agent: schemas.AgentCreate):
    db_agent = models.Agent(**agent.dict())
    db.add(db_agent)
    db.commit()
    db.refresh(db_agent)
    return db_agent

def create_customer(db: Session, customer: schemas.CustomerCreate):
    db_customer = models.Customer(**customer.dict())
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer

def create_campaign1(db: Session, campaign: schemas.CampaignCreate1, company_id: int):
    db_campaign = models.Campaign(
        name=campaign.name,
        email=campaign.email,
        status=campaign.status,
        company_id=company_id
    )
    db.add(db_campaign)
    db.commit()
    db.refresh(db_campaign)
    return db_campaign

def create_campaign(db: Session, campaign_in: schemas.CampaignCreate, file_bytes: bytes, filename: str):
    print("🚀 Creating campaign in DB with data:", campaign_in.dict())

    # Create campaign with gaming_offer
    db_campaign = models.CampaignAgent(
        name=campaign_in.name,
        # email=campaign_in.email,
        company_id=campaign_in.company_id,
        campaign_id=campaign_in.campaign_id,
    )
    db.add(db_campaign)
    db.flush()  # so db_campaign.id is available

    # Create agents dynamically
    agents = []
    for agent_name in campaign_in.agents:
        agent = models.Agent(
            name=agent_name,
            company_id=campaign_in.company_id, 
            campaign_id=campaign_in.campaign_id,        # ✅ assign campaign_id
            campaign_agents_id=db_campaign.id          # ✅ assign campaign_agent_id
        )
        db.add(agent)
        db.flush()
        agents.append(agent)
    db_campaign.agent = agents

    # Handle customers import
    customers = []
    if filename.endswith(".csv"):
        file_like = StringIO(file_bytes.decode("utf-8"))
        reader = csv.DictReader(file_like)
        for row in reader:
            customers.append(
                models.Customer(
                    name=row.get("name"),
                    phone=row.get("phone"),
                    campaign_id=campaign_in.campaign_id,         # ✅ assign campaign_id
                    campaign_agents_id=db_campaign.id           # ✅ assign campaign_agent_id
                )
            )
    elif filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            customers.append(
                models.Customer(
                    name=row_dict.get("name"),
                    phone=row_dict.get("phone"),
                    campaign_id=campaign_in.campaign_id,     # ✅ assign campaign_id
                    campaign_agents_id=db_campaign.id       # ✅ assign campaign_agent_id
                )
            )

    db.add_all(customers)
    db.commit()
    db.refresh(db_campaign)

    print("✅ Saved campaign:", db_campaign.id, db_campaign.gaming_offer)
    print("✅ Agents linked:", [a.name for a in agents])
    print("✅ Customers imported:", len(customers))

    return db_campaign